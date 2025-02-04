from collections import defaultdict
from datetime import date

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from database.db_cmd.cantainers_cmd import get_stored_containers_from_to


async def get_containers_statistic(first_date: date, last_date: date) -> str:
    wb = Workbook()
    del wb["Sheet"]
    ws: Worksheet = wb.create_sheet("Общий список")
    ws_stat: Worksheet = wb.create_sheet("Статистика")
    ws.append(
        ["Номер", "Дата изготовления", "Наименование", "Цвет", "Вес", "Партия пластика", "Артикул крышки", "бригада",
         "Склад", "Комментарий"]
    )

    if last_date < first_date:
        first_date, last_date = last_date, first_date
    containers = await get_stored_containers_from_to(first_date=first_date, last_date=last_date)

    containers_dict = dict()
    batch_dict = defaultdict(lambda: defaultdict(int))
    color_dict = defaultdict(lambda: defaultdict(int))

    total_w = 0

    for container in containers:
        ws.append(container.get_data_list())
        tmp = containers_dict.setdefault(container.name, 0) + 1
        containers_dict[container.name] = tmp
        batch_number = container.batch_number
        color = container.color
        weight = container.weight
        total_w += weight
        batch_dict[batch_number][color] += weight
        color_dict[color][batch_number] += weight

    sheet1 = wb.create_sheet("По партиям")
    sheet1.append(["Партия", "Цвет", "Масса"])

    for batch_number, colors in batch_dict.items():
        for color, total_weight in colors.items():
            sheet1.append([batch_number, color, total_weight])
    sheet1.append([None, None, total_w])

    # Добавляем второй лист (по цветам)
    sheet2 = wb.create_sheet(title="По цветам")
    sheet2.append(["Цвет", "Партия", "Масса"])

    for color, batches in color_dict.items():
        for batch_number, total_weight in batches.items():
            sheet2.append([color, batch_number, total_weight])
    sheet2.append([None, None, total_w])

    ws_stat.append(["Наименование", "Количество"])
    for name, count in sorted(list(containers_dict.items())):
        ws_stat.append([name, count])

    filename = f'Бочки {first_date.strftime("%d.%m.%y")}-{last_date.strftime("%d.%m.%y")} тест.xlsx'
    wb.save(filename)
    return filename
