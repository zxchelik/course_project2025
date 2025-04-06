from datetime import datetime

import openpyxl
from aiofiles import os
from aiogram import Router, F, Bot
from aiogram.filters import Text, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery, FSInputFile, Message
from typing import Callable

from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from src.backend.database.db_cmd.blank_cassetes_cmd import (
    get_tasks,
    AddTaskModel,
    EditTaskModel,
    add_tasks,
    edit_tasks,
    DeleteTaskModel,
)
from src.backend.database.models import Cassette
from src.backend.database.models.blank_cassettes import CassetteType
from src.backend.database.models.cassette import CassetteState
from src.backend.database.models.cassette_group_member import CassetteGroupMemberType
from src.backend.database.session_context import async_session_context
from src.backend.telegram.filters.db_filters import IsAdmin
from src.backend.telegram.states.admin import AddCassetteTask
from src.backend.telegram.utils.auto_size_excel import auto_size_excel

router = Router()


@router.callback_query(IsAdmin(), Text("add_cassette_task"))
async def start(callback: CallbackQuery, state: FSMContext):
    file_name = await generate_file()
    await state.set_state(AddCassetteTask.input_file)
    filename_ = f"ЗАДАЧИ НА КАССЕТЫ {datetime.now().strftime('%y.%m.%d %H-%M')}.xlsx"
    await callback.message.answer_document(
        FSInputFile(file_name, filename=filename_), caption="заполните файл и пришлите его\nИли нажми на /menu"
    )
    await os.remove(file_name)


@router.message(StateFilter(AddCassetteTask.input_file), F.document)
async def process_document(message: Message, state: FSMContext, bot: Bot):
    document = message.document

    if not document.file_name.endswith(".xlsx"):
        await message.answer("Данное расширение файла не поддерживается. Пришли .xlsx файл")
        return

    await bot.download(document, "tmp.xlsx")  # FIXME избавиться от "буферного" файла
    wb = openpyxl.load_workbook("tmp.xlsx")
    add_is_ok, add_results = process_new_tasks(wb)
    edit_is_ok, edit_results, delete_results = process_edit_tasks(wb)
    text = generate_text(add_is_ok, add_results, edit_is_ok, edit_results)

    if add_is_ok and edit_is_ok and (add_results or edit_results or delete_results):
        try:
            await add_tasks(tasks=add_results, customer_id=message.from_user.id)
            await edit_tasks(task_updates=edit_results, task_deletes=delete_results)
            await state.clear()

            text += "\nЧтобы вернуться в меню нажмите /menu"

            file_name = await generate_file()
            await state.set_state(AddCassetteTask.input_file)
            filename_ = f"ЗАДАЧИ НА КАССЕТЫ {datetime.now().strftime('%y.%m.%d %H-%M')}.xlsx"
            await message.answer_document(
                FSInputFile(file_name, filename=filename_), caption="заполните файл и пришлите его\nИли нажми на /menu"
            )
            await os.remove(file_name)
            return
        except Exception as e:
            text = "Произошла непредвиденная ошибка: " + str(e)

    text += "\nЧтобы вернуться в меню нажмите /menu"

    await message.answer(text)


async def generate_file() -> str:
    tasks = await get_tasks()
    wb = Workbook()
    del wb["Sheet"]

    edit_ws: Worksheet = wb.create_sheet("Изменить")
    edit_ws.append(
        [
            "ID",
            "Наименование",
            "Тип",
            "Количество",
            "Новое количество",
            "Приоритет",
            "Новый приоритет",
            "Технический комментарий",
            "Комментарий",
        ]
    )
    for task in tasks:
        edit_ws.append(task.to_list_for_excel())

    add_ws: Worksheet = wb.create_sheet("Добавить")
    add_ws.append(["Наименование", "Количество", "Приоритет", "Тип", "Технический комментарий", "Комментарий"])

    dv_list_each = DataValidation(
        type="list",
        formula1=f'"{",".join(CassetteType.to_list())}"',
        showDropDown=False,
        showErrorMessage=True,
        showInputMessage=True,
        allow_blank=True,
        promptTitle="Тип",
        prompt="Выберите тип изделия",
        error="ВЫБЕРИТЕ ИЗ ПРЕДЛОЖЕННЫХ",
        errorTitle="ОШИБКА ВВОДА",
    )

    add_ws.add_data_validation(dv_list_each)
    dv_list_each.add("D2:D1048576")

    await write_existed_cuted_cassette(wb=wb)

    auto_size_excel(edit_ws)
    file_name = "tmp.xlsx"
    wb.save(file_name)
    return file_name


@async_session_context
async def write_existed_cuted_cassette(session: AsyncSession, wb: Workbook):
    cuted_cassette: Worksheet = wb.create_sheet("Нарезанные кассеты")
    stmt = (
        select(Cassette)
        .where(
            or_(
                Cassette.state == CassetteState.CUT,
                and_(Cassette.state == CassetteState.WELD, Cassette.in_working == True),
            )
        )
        .options(selectinload(Cassette.groups))
    )
    cassettes = (await session.execute(stmt)).scalars().all()
    header = [
        "ID",
        "Дата нарезки",
        "Наименование",
        "Тип",
        "Тех. коммент",
        "В работе",
        "Комментарий",
    ]
    to_list: Callable[[Cassette], list] = lambda c: [
        c.id,
        c.cut_date.strftime("%Y-%m-%d"),
        c.name,
        c.type,
        c.technical_comment,
        (
            ", ".join(
                [
                    i.user.fio
                    for i in c.groups
                    if i.group_type in [CassetteGroupMemberType.WELDER, CassetteGroupMemberType.WELDER_HELPER]
                ]
            )
            if c.in_working == True
            else None
        ),
        c.comment,
    ]
    cuted_cassette.append(header)
    for cassette in cassettes:
        cuted_cassette.append(to_list(cassette))
    auto_size_excel(cuted_cassette)


def process_new_tasks(wb: Workbook) -> tuple[bool, list[AddTaskModel]] | tuple[bool, list[int]]:
    result = []
    bad_results = []
    sheet: Worksheet = wb["Добавить"]
    data = sheet.values
    next(data)
    for i, row in enumerate(data):
        try:
            name, quantity, priority, cassette_type, technical_comment, comment = row
            quantity = int(quantity)
            priority = int(priority)
            result.append(
                AddTaskModel(
                    cassette_name=name,
                    quantity=quantity,
                    priority=priority,
                    type=cassette_type,
                    technical_comment=technical_comment,
                    comment=comment,
                )
            )
        except (ValueError, TypeError):
            bad_results.append(i + 2)

    if bad_results:
        return False, bad_results
    return True, result


def process_edit_tasks(
    wb: Workbook,
) -> tuple[bool, list[EditTaskModel], list[DeleteTaskModel]] | tuple[bool, list[int], None]:
    result = []
    delete_result = []
    bad_results = []
    sheet: Worksheet = wb["Изменить"]
    data = sheet.values
    next(data)
    for i, row in enumerate(data):
        try:
            task_id, _, _, _, new_quantity, _, new_priority, _, _ = row
            if not new_quantity is None:
                new_quantity = int(new_quantity)
                if new_quantity == 0:
                    delete_result.append(DeleteTaskModel(id=task_id))
                    continue
            if not new_priority is None:
                new_priority = int(new_priority)
            result.append(EditTaskModel(id=task_id, new_quantity=new_quantity, new_priority=new_priority))
        except (ValueError, TypeError):
            bad_results.append(i + 2)

    if bad_results:
        return False, bad_results, None
    return True, result, delete_result


def generate_text(
    add_is_ok: bool,
    add_results: list[int] | list[AddTaskModel],
    edit_is_ok: bool,
    edit_results: list[int] | list[EditTaskModel],
) -> str:
    res = "Результат считывания файла:\n\n"

    if not add_results and not edit_results:
        return "Файл не заполнен\nПришлите заполненный файл\n"

    if add_results:
        if add_is_ok:
            res += "Все запросы на дополнение кассет успешно считаны\n"
        else:
            res += "Ошибка при считывании запросов на дополнение кассет\n"
            res += "Строки: " + ", ".join(map(str, add_results)) + "\n"
            res += "Все запросы на дополнение кассет были отклонены\n"

    if edit_results:
        res += "\n"
        if edit_is_ok:
            res += "Все запросы на изменение кассет успешно считаны\n"
        else:
            res += "Ошибка при считывании запросов на изменение кассет\n"
            res += "Строки: " + ", ".join(map(str, edit_results)) + "\n"
            res += "Все запросы на изменение кассет были отклонены\n"

    if not (add_is_ok and edit_is_ok and (add_results or edit_results)):
        res += "\nПришлите исправленный файл\n"

    return res
