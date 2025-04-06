import os
from datetime import datetime

import openpyxl
from aiogram import Router, Bot
from aiogram.filters import Text, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery, FSInputFile, Message
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from src.backend.database.db_cmd.plastic_supply import get_plastic_residue, get_all_colors
from src.backend.database.models import PlasticSupply
from src.backend.database.session_context import async_session_context
from src.backend.telegram.filters.db_filters import IsAdmin
from src.backend.telegram.keyboards.inline import get_inline_kb
from src.backend.telegram.states.admin import GetPlasticResidue
from src.backend.telegram.utils.RAL import load_ral

router = Router()


@router.callback_query(IsAdmin(), Text("get_plastic_residue"))
async def start_get_availability(callback: CallbackQuery, state: FSMContext):
    text = "Выберите"
    kb = get_inline_kb([[("Получить наличие", "get_residue")], [("Поставка", "add_delivery")]])
    await state.set_state(GetPlasticResidue.select_type)
    await callback.message.edit_text(text=text, reply_markup=kb)


async def write_availability(wb: Workbook):
    data = await get_plastic_residue()
    ws: Worksheet = wb.create_sheet("Список")
    ws.append(["Номер цвета", "Осталось"])
    ral = load_ral()
    rgb_to_hex = lambda r, g, b: f"{r:02X}{g:02X}{b:02X}"
    for i, row in enumerate(data):
        ws.append(tuple(row))

        try:
            cn, w = row
            bgcolor = ral[cn]
            r, g, b = bgcolor
            fcolor = "000000" if r * 0.299 + g * 0.587 + b * 0.114 > 186 else "FFFFFF"

            font = Font(color=fcolor)
            fill = PatternFill(start_color=rgb_to_hex(*bgcolor), end_color=rgb_to_hex(*bgcolor), fill_type="solid")

            ws[f"A{i + 2}"].font = font
            ws[f"A{i + 2}"].fill = fill
        except KeyError:
            pass


@router.callback_query(StateFilter(GetPlasticResidue.select_type), Text("get_residue"))
async def get_availability(callback: CallbackQuery, state: FSMContext):
    wb = Workbook()
    del wb["Sheet"]
    await write_availability(wb)
    name = "tmp.xlsx"
    wb.save(name)
    d = datetime.now()
    await callback.message.delete()
    await callback.message.answer_document(FSInputFile(name, filename=f"Остатки {d.strftime('%y.%m.%d %H-%M')}.xlsx"))
    os.remove(name)


@router.callback_query(StateFilter(GetPlasticResidue.select_type), Text("add_delivery"))
async def get_pattern(callback: CallbackQuery, state: FSMContext):
    wb = Workbook()
    del wb["Sheet"]
    data = await get_all_colors()
    ws: Worksheet = wb.create_sheet("Шаблон")
    ws.append(["Номер цвета", "Пришло"])

    ral = load_ral()
    rgb_to_hex = lambda r, g, b: f"{r:02X}{g:02X}{b:02X}"

    for i, row in enumerate(data):
        ws.append([row])

        try:
            cn = row
            bgcolor = ral[cn]
            r, g, b = bgcolor
            fcolor = "000000" if r * 0.299 + g * 0.587 + b * 0.114 > 186 else "FFFFFF"

            font = Font(color=fcolor)
            fill = PatternFill(start_color=rgb_to_hex(*bgcolor), end_color=rgb_to_hex(*bgcolor), fill_type="solid")

            ws[f"A{i + 2}"].font = font
            ws[f"A{i + 2}"].fill = fill
        except KeyError:
            pass

    await write_availability(wb)

    name = "tmp.xlsx"
    wb.save(name)
    await callback.message.delete()
    await callback.message.answer_document(
        FSInputFile(name, filename=f"Шаблон поставок.xlsx"), caption="заполните файл и пришлите его"
    )
    os.remove(name)
    await state.set_state(GetPlasticResidue.get_file)


@router.message(StateFilter(GetPlasticResidue.get_file))
async def read_file(message: Message, bot: Bot, state: FSMContext):
    document = message.document
    if not message.document:
        await message.answer("Пришли файл")
        return

    if not document.file_name.endswith(".xlsx"):
        await message.answer("Данное расширение файла не поддерживается. Пришли .xlsx файл")
        return

    mess = await message.answer("Считываю данные...")

    await bot.download(document, "tmp.xlsx")  # FIXME избавиться от "буферного" файла
    wb = openpyxl.load_workbook("tmp.xlsx")
    sheet: Worksheet = wb.get_sheet_by_name("Шаблон")
    data = sheet.values
    next(data)
    try:

        @async_session_context
        async def add_plastic(session: AsyncSession):
            for row in data:
                if row[1]:
                    session.add(PlasticSupply(color_number=int(row[0]), weight=float(row[1])))
            await session.commit()

        await add_plastic()

        await mess.edit_text("Всё готово✅")
        await state.clear()
    except Exception as e:
        await mess.edit_text(f"Произошла ошибка: {e}")
