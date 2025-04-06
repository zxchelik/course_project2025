import datetime
import os
from typing import Sequence

import openpyxl
from aiogram import Router, Bot, F
from aiogram.filters import Text, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery, FSInputFile, Message
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select, not_, func
from sqlalchemy.ext.asyncio import AsyncSession

from src.backend.database.db_cmd.cantainers_cmd import move_container_with_session
from src.backend.database.models import Container
from src.backend.database.session_context import async_session_context
from src.backend.telegram.filters.db_filters import IsAdmin
from src.backend.telegram.keyboards.inline import get_inline_kb
from src.backend.telegram.states.admin import GetAvailability
from src.backend.telegram.utils.RAL import load_ral

router = Router()


@router.callback_query(IsAdmin(), Text("get_availability"))
async def start_get_availability(callback: CallbackQuery, state: FSMContext):
    text = "Выберите"
    kb = get_inline_kb([[("За всё время", "get_all"), ("В наличии", "on_availability")]])
    await state.set_state(GetAvailability.select_type)
    await callback.message.edit_text(text=text, reply_markup=kb)


@router.callback_query(StateFilter(GetAvailability.select_type))
async def send_pattern(callback: CallbackQuery, state: FSMContext):
    file_type = callback.data

    await state.set_state(GetAvailability.get_file)
    msg = callback.message
    await msg.edit_text("Генерирую файл...")

    @async_session_context
    async def get_containers(session: AsyncSession) -> Sequence[Container]:
        stmt = (
            select(Container)
            .order_by(Container.storage, Container.name, Container.number)
            .where(not_(Container.number.contains("-")))
            if file_type == "get_all"
            else select(Container)
            .order_by(Container.storage, Container.name, Container.number)
            .where(not_(Container.number.contains("-")))
            .where(Container.storage != "Отгружено")
            .where(Container.storage != "Н/Д")
        )
        containers_ = (await session.scalars(stmt)).all()
        return containers_

    @async_session_context
    async def get_statistic(session: AsyncSession):
        stmt = (
            select(Container.name, func.count().label("count"))
            .order_by(Container.name)
            .group_by(Container.name)
            .where(not_(Container.number.contains("-")))
            .where(Container.storage != "Отгружено")
            .where(Container.storage != "Н/Д")
        )
        return (await session.execute(stmt)).all()

    containers = await get_containers()
    wb = Workbook()
    del wb["Sheet"]

    if file_type != "get_all":
        ws_stat: Worksheet = wb.create_sheet("Статистика")
        ws_stat.append(["Наименование", "Кол-во"])
        ws_stat.column_dimensions["A"].width = 15
        ws_stat.column_dimensions["B"].width = 8
        [ws_stat.append([*i]) for i in (await get_statistic())]

    ws: Worksheet = wb.create_sheet("Список")
    ws.append(["Наименование", "Номер", "Цвет", "Место", "Новое место", "Основание перемещения", "Комментарий"])

    ral = load_ral()
    rgb_to_hex = lambda r, g, b: f"{r:02X}{g:02X}{b:02X}"

    column_sizes = {"A": 15, "B": 10, "C": 7, "D": 8, "E": 11, "F": 22, "G": 17}

    for column, size in column_sizes.items():
        ws.column_dimensions[column].width = size

    for i, c in enumerate(containers):
        ws.append([c.name, c.number, c.color, c.storage, None, None, c.comments])

        try:
            cn = c.color
            bgcolor = ral[cn]
            r, g, b = bgcolor
            fcolor = "000000" if r * 0.299 + g * 0.587 + b * 0.114 > 186 else "FFFFFF"

            font = Font(color=fcolor)
            fill = PatternFill(start_color=rgb_to_hex(*bgcolor), end_color=rgb_to_hex(*bgcolor), fill_type="solid")

            ws[f"C{i + 2}"].font = font
            ws[f"C{i + 2}"].fill = fill
        except KeyError:
            pass

    await msg.delete()
    name = "tmp.xlsx"
    wb.save(name)
    d = datetime.datetime.now()
    filename_ = f"{'ВСЕ ' if file_type == 'get_all' else ''}БОЧКИ {d.strftime('%y.%m.%d %H-%M')}.xlsx"
    msg = await msg.answer_document(
        FSInputFile(name, filename=filename_), caption="заполните файл и пришлите его\nИли нажми на /menu"
    )
    os.remove(name)
    await state.update_data(msg=msg)


@router.message(StateFilter(GetAvailability.get_file), F.document)
async def get_file(message: Message, state: FSMContext, bot: Bot):
    document = message.document
    if not message.document:
        await message.answer("Пришли файл")
        return

    if not document.file_name.endswith(".xlsx"):
        await message.answer("Данное расширение файла не поддерживается. Пришли .xlsx файл")
        return

    mess = await message.answer("Считываю данные...")
    user_id = message.from_user.id

    await bot.download(document, "tmp.xlsx")  # FIXME избавиться от "буферного" файла
    wb = openpyxl.load_workbook("tmp.xlsx")
    sheet: Worksheet = wb["Список"]
    data = sheet.values

    without_rfm = []
    next(data)
    try:
        for row in data:
            number, to_storage, reasons_for_moving = row[1], row[4], row[5]
            if to_storage:
                if reasons_for_moving:
                    try:
                        await move_container_with_session(
                            number=number, to_storage=to_storage, user_id=user_id, reasons_for_moving=reasons_for_moving
                        )
                    except ValueError:
                        raise
                else:
                    without_rfm.append(number)
        newline = "\n"
        text = (
            f"""Ошибка!
Не указано основание перемещения для бочек с номером:
{newline.join(without_rfm)}

Повторите попытку"""
            if without_rfm
            else "Всё готово✅"
        )

        await mess.edit_text(text)
        await state.clear()
    except Exception as e:
        await mess.edit_text(f"Произошла ошибка: {e}")
